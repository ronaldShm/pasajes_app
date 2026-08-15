"""Exportador y gestor de archivos Excel."""
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from config import EXCEL_PATH, COLUMNAS_EXCEL


class ExcelExporter:
    def __init__(self):
        self.excel_path = EXCEL_PATH
        self.wb = None
        self.ws = None
        self._load_or_create()

    def _load_or_create(self):
        if self.excel_path.exists():
            try:
                self.wb = load_workbook(str(self.excel_path))
                self.ws = self.wb.active
                return
            except Exception:
                pass
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Pasajes"
        self._create_headers()

    def _create_headers(self):
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="2E86AB", end_color="2E86AB", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(COLUMNAS_EXCEL, 1):
            cell = self.ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        col_widths = {
            1: 18, 2: 12, 3: 45, 4: 15, 5: 30, 6: 10,
            7: 14, 8: 18, 9: 12, 10: 12, 11: 14,
            12: 15, 13: 22, 14: 20, 15: 15, 16: 30,
        }
        for col, width in col_widths.items():
            self.ws.column_dimensions[self._get_column_letter(col)].width = width

    def _get_column_letter(self, col: int) -> str:
        result = ""
        while col > 0:
            col, remainder = divmod(col - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def add_record(self, data: dict):
        row = self.ws.max_row + 1
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        values = [
            data.get("fecha_registro", ""),
            data.get("aerolinea", ""),
            data.get("pasajeros", ""),
            data.get("cantidad_pasajeros", 1),
            data.get("ticket", ""),
            data.get("reserva", ""),
            data.get("fecha_emision", ""),
            data.get("vuelo", ""),
            data.get("origen", ""),
            data.get("destino", ""),
            data.get("fecha_vuelo", ""),
            data.get("total_pagado", ""),
            data.get("forma_pago", ""),
            data.get("solicitado_por", ""),
            data.get("ceco", ""),
            data.get("archivo_origen", ""),
        ]
        for col_idx, value in enumerate(values, 1):
            cell = self.ws.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    def save(self):
        if self.wb:
            self.wb.save(str(self.excel_path))

    def clear_data(self):
        if self.ws and self.ws.max_row > 1:
            self.ws.delete_rows(2, self.ws.max_row - 1)

    def export_from_records(self, records: list[dict]):
        self.clear_data()
        for record in records:
            self.add_record(record)
        self.save()

    def get_all_records(self) -> list[dict]:
        if not self.ws or self.ws.max_row <= 1:
            return []
        records = []
        headers = [cell.value for cell in self.ws[1]]
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            records.append(record)
        return records

    def export_report_to_excel(self, report_data: dict, filename: str = None):
        if not filename:
            filename = f"reporte_pasajes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        report_path = Path.home() / "Downloads" / filename
        wb = Workbook()
        
        # Resumen General
        ws_summary = wb.active
        ws_summary.title = "Resumen General"
        ws_summary.append(["Métrica", "Valor"])
        summary = report_data.get("resumen", {})
        ws_summary.append(["Total Pasajes", summary.get("total_pasajes", 0)])
        ws_summary.append(["Total Gastado", f"${summary.get('total_gastado', 0):,.0f}"])
        ws_summary.append(["Promedio por Pasaje", f"${summary.get('promedio', 0):,.0f}"])
        ws_summary.append(["Pasajeros Únicos", summary.get("total_pasajeros", 0)])
        
        # Gasto por CECO
        ws_ceco = wb.create_sheet("Gasto por CECO")
        ws_ceco.append(["CECO", "Pasajes", "Total", "Promedio"])
        for item in report_data.get("ceco", []):
            ws_ceco.append([
                item.get("ceco", ""),
                item.get("pasajes", 0),
                f"${item.get('total', 0):,.0f}",
                f"${item.get('promedio', 0):,.0f}"
            ])
        
        # Gasto por Solicitante
        ws_solicitante = wb.create_sheet("Gasto por Solicitante")
        ws_solicitante.append(["Solicitante", "Pasajes", "Total", "Promedio"])
        for item in report_data.get("solicitante", []):
            ws_solicitante.append([
                item.get("solicitado_por", ""),
                item.get("pasajes", 0),
                f"${item.get('total', 0):,.0f}",
                f"${item.get('promedio', 0):,.0f}"
            ])
        
        # Gasto por Aerolínea
        ws_aerolinea = wb.create_sheet("Gasto por Aerolínea")
        ws_aerolinea.append(["Aerolínea", "Pasajes", "Total", "Promedio", "Mínimo", "Máximo"])
        for item in report_data.get("aerolinea", []):
            ws_aerolinea.append([
                item.get("aerolinea", ""),
                item.get("pasajes", 0),
                f"${item.get('total', 0):,.0f}",
                f"${item.get('promedio', 0):,.0f}",
                f"${item.get('minimo', 0):,.0f}",
                f"${item.get('maximo', 0):,.0f}"
            ])
        
        # Top Rutas
        ws_rutas = wb.create_sheet("Top Rutas")
        ws_rutas.append(["Origen", "Destino", "Viajes", "Total", "Promedio"])
        for item in report_data.get("rutas", []):
            ws_rutas.append([
                item.get("origen", ""),
                item.get("destino", ""),
                item.get("viajes", 0),
                f"${item.get('total', 0):,.0f}",
                f"${item.get('promedio', 0):,.0f}"
            ])
        
        # Top Pasajeros
        ws_pasajeros = wb.create_sheet("Top Pasajeros")
        ws_pasajeros.append(["Pasajero", "Viajes", "Total"])
        for item in report_data.get("pasajeros", []):
            ws_pasajeros.append([
                item.get("pasajeros", ""),
                item.get("viajes", 0),
                f"${item.get('total', 0):,.0f}"
            ])
        
        wb.save(str(report_path))
        return report_path
